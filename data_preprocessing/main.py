import pandas as pd
import numpy as np
from pandas import DataFrame
import os
from datetime import datetime, date
from math import radians, cos, sin, asin, sqrt
import os
import openpyxl
import codecs
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


# 计算第一类异常点
def stopPoint(file):
    data = pd.read_excel(file, header=None, index_col=None)
    # print(data)
    data_array = np.array(data)
    data_list = data_array.tolist()
    # print(data_list)
    neg_num = 0
    neg_List = []
    for i in range(len(data_list)):
        if i == 0:
            continue
        else:
            if (data_list[i][3] == data_list[i - 1][3] and data_list[i][4] == data_list[i - 1][4]
                    and data_list[i][7] == data_list[i - 1][7] and data_list[i][8] == data_list[i - 1][8]
                    and data_list[i - 1][7] > 2):
                neg_List.append(i)
                neg_num = neg_num + 1

    return neg_List, neg_num


# 计算时间差
def timeCalculate(t1, t2):
    t1_struct = (datetime.strptime(t1, "%Y-%m-%d %H:%M:%S"))
    t2_struct = (datetime.strptime(t2, "%Y-%m-%d %H:%M:%S"))
    total_seconds = (t2_struct - t1_struct).total_seconds()
    return total_seconds


# 计算距离
def distanceCalculate(x1, y1, x2, y2):
    x1, y1, x2, y2 = map(radians, [float(x1), float(y1), float(x2), float(y2)])
    dx = x2 - x1
    dy = y2 - y1
    a = sin(dy / 2) ** 2 + cos(y1) * cos(y2) * sin(dx / 2) ** 2
    c = 2 * asin(sqrt(a))
    r = 6371
    return c * r * 1000


# 寻找第二类异常点
def neg2Point(file):
    data = pd.read_excel(file, header=None, index_col=None)
    data_array = np.array(data)
    data_list = data_array.tolist()
    neg_num = 0
    neg_List = []
    check = []

    for i in range(len(data_list)):
        if i == 0:
            check.append(True)
            continue
        else:
            j = i - 1
            while not check[j]:
                j = j - 1
            t1 = data_list[j][0] + ' ' + data_list[j][1]
            t2 = data_list[i][0] + ' ' + data_list[i][1]
            t = timeCalculate(t1, t2)
            dis = distanceCalculate(data_list[j][3], data_list[j][4], data_list[i][3], data_list[i][4])
            v = dis / t
            if v*1.94 <= 2*max(data_list[i][7], data_list[j][7]):
                check.append(True)
            elif max(data_list[i][7], data_list[j][7]) <= 0.2:
                check.append(True)
            else:
                check.append(False)
                neg_num = neg_num + 1
                neg_List.append(i)
    # print(file)
    # print(neg_List)
    print(neg_num)
    return neg_List, neg_num


# txt文件转excel
def txt_to_xlsx(filename1, filename2, outfile, sheetname):
    fr = codecs.open(filename1, 'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    #neg_point = stopPoint(filename2)[0]
    neg_point = neg2Point(filename2)[0]
    # ws = wb.create_sheet()
    cnt = 0
    ws.title = sheetname
    row = 0
    realrow = 0
    for line in fr:

        realrow += 1
        line = line.strip()
        line = line.split(' ')
        col = 0
        if cnt < len(neg_point) and realrow == neg_point[cnt] + 1:
            cnt += 1
            continue
        else:
            row += 1
            for j in range(len(line)):
                col += 1
                # print (line[j])
                if (col != 1) and (col != 2) and (col != 10):
                    ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col))).data_type = "int"
                else:
                    ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))

    wb.save(outfile)


if __name__ == '__main__':
    # inputdir_path1 = 'C:\\Users\\asus\\Desktop\\raw_format'
    # inputdir_path2 = 'C:\\Users\\asus\\Desktop\\raw_format_xlsx'
    # fileList1 = os.listdir(inputdir_path1)
    # fileList2 = os.listdir(inputdir_path2)
    # outputdir_path = 'C:\\Users\\asus\\Desktop\\DeleteStopPoint_xlsx'
    inputdir_path1 = 'C:\\Users\\asus\\Desktop\\DeleteStopPoint'
    inputdir_path2 = 'C:\\Users\\asus\\Desktop\\DeleteStopPoint_xlsx'
    fileList1 = os.listdir(inputdir_path1)
    fileList2 = os.listdir(inputdir_path2)
    outputdir_path = 'C:\\Users\\asus\\Desktop\\DeleteFlyPoint_xlsx'
    for i in range(len(fileList1)):
        inputfileTxt1 = os.path.join(inputdir_path1, fileList1[i])
        inputfileTxt2 = os.path.join(inputdir_path2, fileList2[i])
        outfileExcel = os.path.join(outputdir_path, fileList1[i]).replace('.txt', '.xlsx')
        sheetname = fileList1[i].replace(".txt", '')
        print(sheetname)
        txt_to_xlsx(inputfileTxt1, inputfileTxt2,outfileExcel, sheetname)
