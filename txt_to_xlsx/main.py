
import os
import openpyxl
import codecs
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


def txt_to_xlsx(filename, outfile, sheetname):
    fr = codecs.open(filename, 'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    # ws = wb.create_sheet()
    ws.title = sheetname
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split(' ')
        col = 0

        for j in range(len(line)):
            col += 1
            # print (line[j])
            if j==0:
                if (col != 1) and (col != 2) and (col != 10):
                    ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col))).data_type = "int"
                else:
                    ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))

    wb.save(outfile)


# # 读取xlsx内容
# def read_xlsx(filename):
#     # 载入文件
#     wb = openpyxl.load_workbook(filename)
#     # 获取Sheet1工作表
#     ws = wb.get_sheet_by_name('Sheet1')
#     # 按行读取
#     for row in ws.rows:
#         for cell in row:
#             print(cell.value)
#     # 按列读
#     for col in ws.columns:
#         for cell in col:
#             print(cell.value)


if __name__ == '__main__':
    inputdir_path = 'C:\\Users\\asus\\Desktop\\raw_format'
    outputdir_path = 'C:\\Users\\asus\\Desktop\\raw_format_xlsx'
    fileList = os.listdir(inputdir_path)
    print(fileList)
    for name in fileList:
        inputfileTxt = os.path.join(inputdir_path, name)
        outfileExcel = os.path.join(outputdir_path, name).replace('.txt', '.xlsx')
        sheetname = name.replace(".txt", '')
        print(sheetname)
        txt_to_xlsx(inputfileTxt, outfileExcel, sheetname)
