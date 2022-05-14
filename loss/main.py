import os
import openpyxl
import codecs
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import datetime
from openpyxl import Workbook
def timeCalculate(t1, t2):
    t1_struct = (datetime.datetime.strptime(t1, "%Y-%m-%d %H:%M:%S"))
    t2_struct = (datetime.datetime.strptime(t2, "%Y-%m-%d %H:%M:%S"))
    total_seconds = (t2_struct - t1_struct).total_seconds()
    return total_seconds
def processLose(filename, outfile, sheetname):
    df = pd.read_csv(filename, header=None, sep=" ")
    data_array = np.array(df)
    data_list = data_array.tolist()
    fr = codecs.open(filename, 'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    # ws = wb.create_sheet()
    ws.title = sheetname
    row = 0
    real_row = 0
    for line in fr:
        row += 1
        if row == 1:
            real_row += 1
            line = line.strip()
            line = line.split(' ')
            col = 0
            for j in range(len(line)):
                col += 1
                # print (line[j])
                if (col != 1) and (col != 2) and (col != 10):
                    ws.cell(column=col, row=real_row, value=line[j].format(get_column_letter(col))).data_type = "int"
                else:
                    ws.cell(column=col, row=real_row, value=line[j].format(get_column_letter(col)))
        else:
            t1 = data_list[row - 1][0] + ' ' + data_list[row - 1][1]
            t2 = data_list[row - 2][0] + ' ' + data_list[row - 2][1]
            t = timeCalculate(t2, t1)
            if t <= 360 or t >= 7200:
                real_row += 1
                line = line.strip()
                line = line.split(' ')
                col = 0
                for j in range(len(line)):
                    col += 1
                    # print (line[j])
                    if (col != 1) and (col != 2) and (col != 10):
                        ws.cell(column=col, row=real_row,value=line[j].format(get_column_letter(col))).data_type = "int"
                    else:
                        ws.cell(column=col, row=real_row, value=line[j].format(get_column_letter(col)))
            else:
                n = (int)(t / 360)
                for i in range(1, n + 1):
                    real_row += 1
                    ws.cell(column=1, row=real_row,value=data_list[row - 2][0])
                    ws.cell(column=2, row=real_row,value=((datetime.datetime.strptime(data_list[row - 2][1], "%H:%M:%S")+ datetime.timedelta(seconds=i * 360)).strftime("%H:%M:%S")))
                    ws.cell(column=3, row=real_row,value=data_list[row - 2][2])
                    ws.cell(column=4, row=real_row,value=(data_list[row - 1][3] -data_list[row - 2][3])/ t * 360 * i + data_list[row - 2][3])
                    ws.cell(column=5, row=real_row,value=(data_list[row - 1][4] -data_list[row - 2][4])/ t * 360 * i + data_list[row - 2][4])
                col = 0
                line = line.strip()
                line = line.split(' ')
                real_row += 1
                for j in range(len(line)):
                    col += 1
                    # print (line[j])
                    if (col != 1) and (col != 2) and (col != 10):
                        ws.cell(column=col, row=real_row,value=line[j].format(get_column_letter(col))).data_type = "int"
                    else:
                        ws.cell(column=col, row=real_row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)
if __name__ == '__main__':
    inputdir_path = 'C:\\Users\\asus\\Desktop\\DeleteFlyPoint'
    outputdir_path = 'C:\\Users\\asus\\Desktop\\loss_xlsx'
    fileList = os.listdir(inputdir_path)
    print(fileList)
    for name in fileList:
        inputfileTxt = os.path.join(inputdir_path, name)
        outfileExcel = os.path.join(outputdir_path, name).replace('.txt', '.xlsx')
        sheetname = name.replace(".txt", '.xlsx')
        print(sheetname)
        processLose(inputfileTxt, outfileExcel, sheetname)